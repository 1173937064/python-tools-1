import openpyxl

def Excel():
    wb = openpyxl.load_workbook('./file/资料表.xlsx')       #导入Excel文件
    sheet = wb.worksheets[0]                                    #选择sheet
    wb1=openpyxl.load_workbook('./file/表.xlsx')  #当前文件夹及其下属文件夹用‘\\’
    sheet2 = wb1.worksheets[0]

    flag=0                                  #用于计数
    for j in sheet.values:       #iter_rows(min_row=2,max_row=794).values:
        sheet2['E4'] = j[0]
        sheet2['L4'] = j[3]
        sheet2['E5'] = j[1]
        sheet2['L5'] = j[4]
        sheet2['E6'] = j[2]
        sheet2['E7'] = j[4]
        wb1.save('./final/%d_.xlsx'%(flag))
        flag+=1
    return 0
        

Excel()


#常用函数
'''
1.wb = openpyxl.load_workbook('./xx/xx.xlsx')   #读取文件
  1.1 wb1 = Workbook()                        #创建新的工作簿
2.sheet = wb.worksheets[0]                   #选择sheet  0也可是xxx
  2.1 ws1 = wb.create_sheet('xxx')          #指定sheet
  2.2 sheet.values                          #代表sheet里的值，类似二元数组
  2.3 sheet['A1']                           #选定单元格
  2.4 sheet.cell(1,1)                       #先行后列，索引下标
  2.5 sheet['A:B']   sheet[5:10]    sheet['A3:B9']     #多行
3.wb.save('xx.xlsx')                        #保存文件
4.单元格属性
  4.1 cell = ws['A1']
  4.2 cell.col_idx                          #单元格索引
  4.3 cell.colum
  4.4 cell.row                              #单元格行索引
  4.5 cell.colum_letter                     #单元格列名
  4.6 cell.coordinate                       #单元格的坐标

5.赋值
    5.1 ws['A1'] = 20   ws.cell(2,2).value = 20   #使用cell只能给value赋值
    5.2 ws.append([1,2,3])      #增加一行
    
6.最大行、最大列
    5.1 ws.max_column    ws.max_row

7.删除行或列
    7.1 ws.delete_cols(1)
    7.2 ws.delete_rows(2)

8.转pands
    import pandas as pd
    df = pd..DataFrame(ws.values)
    df
    
    pandas 转  ws
    for i in df.values:
    ws.append(i.tolist())
    for i in ws.rows:
    for j in i:
        print(j,j.value,end=',')
    print('')
9.合并单元格
    9.1 ws.merge_cells("A1:B1")
    ws.merge_cells(start_column=3,end_column=6,start_row=2,end_row=3)
    已存在的合并单元格
    ws.merged_cells

10.字体
ws.cell(5,3).value='哈哈哈'
ws.cell(5,3).font = Font(name='仿宋',size=12,color=Color(index=0),b=True,i=True)

# size   sz  字体大小
# b bold  是否粗体
# i italic  是否斜体
# name family  字体样式

11.边框
Side(style='thin',color=Color(index=0))

# style可选项
style = ('dashDot','dashDotDot', 'dashed','dotted',
'double','hair', 'medium', 'mediumDashDot', 'mediumDashDotDot',
'mediumDashed', 'slantDashDot', 'thick', 'thin')
#  'medium' 中粗
#  'thin'  细
#  'thick'  粗
#  'dashed'  虚线
#  'dotted'  点线



颜色
Color(index=0) # 根据索引进行填充
# 
Color(rgb='00000000') # 根据rgb值进行填充
# index 
COLOR_INDEX = (
    '00000000', '00FFFFFF', '00FF0000', '0000FF00', '000000FF', #0-4
    '00FFFF00', '00FF00FF', '0000FFFF', '00000000', '00FFFFFF', #5-9
    '00FF0000', '0000FF00', '000000FF', '00FFFF00', '00FF00FF', #10-14
    '0000FFFF', '00800000', '00008000', '00000080', '00808000', #15-19
    '00800080', '00008080', '00C0C0C0', '00808080', '009999FF', #20-24
    '00993366', '00FFFFCC', '00CCFFFF', '00660066', '00FF8080', #25-29
    '000066CC', '00CCCCFF', '00000080', '00FF00FF', '00FFFF00', #30-34
    '0000FFFF', '00800080', '00800000', '00008080', '000000FF', #35-39
    '0000CCFF', '00CCFFFF', '00CCFFCC', '00FFFF99', '0099CCFF', #40-44
    '00FF99CC', '00CC99FF', '00FFCC99', '003366FF', '0033CCCC', #45-49
    '0099CC00', '00FFCC00', '00FF9900', '00FF6600', '00666699', #50-54
    '00969696', '00003366', '00339966', '00003300', '00333300', #55-59
    '00993300', '00993366', '00333399', '00333333',  #60-63
)
BLACK = COLOR_INDEX[0]
WHITE = COLOR_INDEX[1]
RED = COLOR_INDEX[2]
DARKRED = COLOR_INDEX[8]
BLUE = COLOR_INDEX[4]
DARKBLUE = COLOR_INDEX[12]
GREEN = COLOR_INDEX[3]
DARKGREEN = COLOR_INDEX[9]
YELLOW = COLOR_INDEX[5]
DARKYELLOW = COLOR_INDEX[19]


  
'''




